"""Excel 数据采集器 - 从 Excel 文件采集数据"""

import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import load_workbook

from src.collector.semantic_parser import SemanticParser
from src.utils.entity_resolver import EntityResolver


class ExcelCollector:
    """Excel 数据采集器"""

    # 默认处理的 Sheet 名称
    DEFAULT_SHEETS = ["汇总表", "综合分析报表", "周数据汇总表"]

    def __init__(self, dict_dir: str = "data/dictionaries"):
        self.resolver = EntityResolver(dict_dir)
        self.parser = SemanticParser(self.resolver)

    def collect(
        self,
        file_path: str,
        sheets: Optional[List[str]] = None,
        year: Optional[int] = None,
        week: Optional[int] = None
    ) -> Tuple[Dict, List[Dict]]:
        """
        从 Excel 文件采集数据

        Args:
            file_path: Excel 文件路径
            sheets: 要处理的 Sheet 列表（默认全部）
            year: 年份（用于元数据）
            week: 周数（用于元数据）

        Returns:
            (采集结果, 错误列表)
        """
        file_path = Path(file_path)
        if not file_path.exists():
            return {}, [{"level": "ERROR", "message": f"文件不存在: {file_path}"}]

        # 加载 Excel
        try:
            wb = load_workbook(file_path, data_only=True)
        except Exception as e:
            return {}, [{"level": "ERROR", "message": f"无法打开文件: {e}"}]

        # 确定 Sheet 列表
        if sheets is None:
            sheets = [s for s in wb.sheetnames if s in self.DEFAULT_SHEETS]
            if not sheets:
                sheets = wb.sheetnames[:3]  # 默认取前 3 个

        # 初始化结果
        result = {
            "meta": {
                "year": year or self._extract_year(file_path.name),
                "week": week or self._extract_week(file_path.name),
                "extracted_at": datetime.now().isoformat(),
                "source_file": file_path.name,
                "sheets_processed": sheets
            },
            "organizations": {},
            "validation_report": {
                "status": "pending",
                "errors": [],
                "warnings": [],
                "coverage": 0
            }
        }

        all_errors = []

        # 逐 Sheet 处理（容错：部分失败不影响整体）
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                all_errors.append({
                    "level": "WARN",
                    "message": f"Sheet 不存在: {sheet_name}",
                    "sheet": sheet_name
                })
                continue

            ws = wb[sheet_name]
            sheet_data, sheet_errors = self.parser.parse_sheet(ws, sheet_name)

            # 合并数据
            for org_name, org_data in sheet_data.get("organizations", {}).items():
                if org_name not in result["organizations"]:
                    result["organizations"][org_name] = org_data
                else:
                    # 合并指标（优先保留已有数据）
                    for energy_type, metrics in org_data.get("metrics", {}).items():
                        if energy_type not in result["organizations"][org_name]["metrics"]:
                            result["organizations"][org_name]["metrics"][energy_type] = metrics

            all_errors.extend(sheet_errors)

        # 计算覆盖率
        coverage = self._calculate_coverage(result)
        result["validation_report"]["coverage"] = coverage
        result["validation_report"]["errors"] = all_errors
        result["validation_report"]["status"] = "pass" if coverage >= 95 else "warning"

        wb.close()
        return result, all_errors

    def _extract_year(self, filename: str) -> Optional[int]:
        """从文件名提取年份"""
        import re
        match = re.search(r"(\d{4})年", filename)
        if match:
            return int(match.group(1))
        return None

    def _extract_week(self, filename: str) -> Optional[int]:
        """从文件名提取周数"""
        import re
        match = re.search(r"第(\d+)周", filename)
        if match:
            return int(match.group(1))
        return None

    def _calculate_coverage(self, result: Dict) -> float:
        """计算数据覆盖率"""
        org_count = len(result.get("organizations", {}))
        if org_count == 0:
            return 0

        # 预期指标数（每个组织应有合计的电量、电价、电费）
        expected_fields_per_org = 3  # 电量、电价、电费
        expected_total = org_count * expected_fields_per_org

        # 实际有效数据数
        actual_count = 0
        for org_data in result.get("organizations", {}).values():
            metrics = org_data.get("metrics", {})
            for energy_type, energy_metrics in metrics.items():
                for metric_name, metric_data in energy_metrics.items():
                    if metric_data.get("value") is not None:
                        actual_count += 1

        coverage = (actual_count / expected_total) * 100 if expected_total > 0 else 0
        return round(coverage, 2)

    def collect_batch(
        self,
        file_paths: List[str],
        year: Optional[int] = None,
        week: Optional[int] = None
    ) -> Tuple[Dict, List[Dict]]:
        """
        批量采集多个文件

        Args:
            file_paths: 文件路径列表
            year: 年份
            week: 周数

        Returns:
            (合并结果, 合并错误)
        """
        merged_result = {
            "meta": {
                "year": year,
                "week": week,
                "extracted_at": datetime.now().isoformat(),
                "source_files": [Path(f).name for f in file_paths]
            },
            "organizations": {}
        }
        all_errors = []

        for file_path in file_paths:
            file_result, file_errors = self.collect(file_path, year=year, week=week)

            # 合并组织数据
            for org_name, org_data in file_result.get("organizations", {}).items():
                if org_name not in merged_result["organizations"]:
                    merged_result["organizations"][org_name] = org_data
                else:
                    # 合并指标
                    for energy_type, metrics in org_data.get("metrics", {}).items():
                        if energy_type not in merged_result["organizations"][org_name]["metrics"]:
                            merged_result["organizations"][org_name]["metrics"][energy_type] = metrics

            all_errors.extend(file_errors)

        return merged_result, all_errors

    def save_json(
        self,
        result: Dict,
        output_path: str,
        schema_version: str = "1.0.0"
    ) -> str:
        """
        保存结果到 JSON 文件

        Args:
            result: 采集结果
            output_path: 输出路径
            schema_version: Schema 版本

        Returns:
            保存的文件路径
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        # 添加 schema 版本
        result["schema_version"] = schema_version

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)

        return str(output_path)


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description="Excel 数据采集")
    parser.add_argument("--input", required=True, help="Excel 文件路径")
    parser.add_argument("--output", required=True, help="JSON 输出路径")
    parser.add_argument("--year", type=int, help="年份")
    parser.add_argument("--week", type=int, help="周数")
    parser.add_argument("--sheets", nargs="+", help="Sheet 名称列表")

    args = parser.parse_args()

    collector = ExcelCollector()
    result, errors = collector.collect(
        args.input,
        sheets=args.sheets,
        year=args.year,
        week=args.week
    )

    # 输出错误信息
    for err in errors:
        level = err.get("level", "INFO")
        msg = err.get("message", "")
        print(f"[{level}] {msg}")

    # 保存结果
    output_file = collector.save_json(result, args.output)
    print(f"数据已保存到: {output_file}")
    print(f"覆盖率: {result['validation_report']['coverage']}%")


if __name__ == "__main__":
    main()