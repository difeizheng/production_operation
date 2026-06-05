"""汇总表补充采集器 - 从汇总表采集现货价格和原因描述

从 2026年第XX周周数据汇总表.xlsx 采集两类补充数据：
    1. 现货市场均价（营销区域周现货价格信息填报表）
    2. 原因描述文本（国内数据填报表 H列）

遵循与 AnalysisCollector 一致的接口：返回 (data_dict, error_list)。
"""

import logging
import re
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# 现货市场需要采集的地区（与真实文档 Table 2 一致）
SPOT_REGIONS: List[str] = [
    "广东", "山西", "山东", "甘肃", "蒙西", "湖北", "浙江", "陕西",
    "西班牙", "巴西",
]

# 国内数据填报表中原因文本的位置
# (行号, 字段名)
REASON_ROWS: List[Tuple[int, str]] = [
    (47, "yoy_summary"),       # 汇总同比原因
    (71, "wow_summary"),       # 汇总环比原因
    (29, "yoy_changjiang"),    # 长江电力同比原因
    (53, "wow_changjiang"),    # 长江电力环比原因
]

REASON_COL: int = 8  # H列


class SummaryCollector:
    """汇总表补充采集器

    从 周数据汇总表.xlsx 采集现货价格和原因描述等补充数据。
    """

    def __init__(self) -> None:
        self._errors: List[Dict[str, Any]] = []

    def collect(
        self,
        file_path: str,
        year: Optional[int] = None,
        week: Optional[int] = None,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """从汇总表采集补充数据。

        Args:
            file_path: 汇总表.xlsx 文件路径
            year: 年份（用于元数据）
            week: 周数（用于元数据）

        Returns:
            (采集结果字典, 错误列表)
        """
        self._errors = []
        file_path = Path(file_path)

        if not file_path.exists():
            return {}, [{"level": "ERROR", "message": f"文件不存在: {file_path}"}]

        try:
            wb = load_workbook(str(file_path), data_only=True)
        except Exception as e:
            return {}, [{"level": "ERROR", "message": f"无法打开文件: {e}"}]

        result: Dict[str, Any] = {}

        # 1. 采集现货市场均价
        result["spot_prices"] = self._collect_spot_prices(wb)

        # 2. 采集原因描述文本
        result["reasons"] = self._collect_reasons(wb)

        # 元数据
        result["meta"] = {
            "year": year or self._extract_year(file_path.name),
            "week": week or self._extract_week(file_path.name),
            "extracted_at": datetime.now().isoformat(),
            "source_file": file_path.name,
            "collector": "SummaryCollector",
        }

        wb.close()
        return result, self._errors

    # ========================================================================
    # 现货市场均价采集
    # ========================================================================

    def _collect_spot_prices(self, wb: Any) -> Dict[str, Any]:
        """从营销区域周现货价格信息填报表采集现货均价。

        每个省份有4行（风电/光伏/新能源/全类型）。
        优先取"全类型"行，若全类型无效则取"新能源"行，最后取任意有效行。

        Returns:
            {regions: [...], data: {地区: {avg, yoy, wow, ...}}}
        """
        sheet_name = "营销区域周现货价格信息填报表"
        if sheet_name not in wb.sheetnames:
            self._errors.append({
                "level": "WARNING",
                "message": f"Sheet '{sheet_name}' 不存在，跳过现货数据",
            })
            return {"regions": SPOT_REGIONS, "data": {}}

        ws = wb[sheet_name]

        # 先按省份收集所有候选行，再按优先级选取
        candidates: Dict[str, List[Tuple[str, int]]] = {}
        current_province: Optional[str] = None

        for row_idx in range(2, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value

            if a_val is not None and b_val is not None and isinstance(b_val, str):
                current_province = b_val.strip()

            if current_province is None:
                continue
            if current_province not in SPOT_REGIONS:
                continue

            e_val = ws.cell(row=row_idx, column=5).value   # E: 电源类型
            i_val = ws.cell(row=row_idx, column=9).value   # I: 本周均价
            k_val = ws.cell(row=row_idx, column=11).value   # K: 同比

            if not self._is_valid_spot_row(i_val, k_val):
                continue

            type_name = str(e_val).strip() if e_val else ""
            if current_province not in candidates:
                candidates[current_province] = []
            candidates[current_province].append((type_name, row_idx))

        # 按优先级选取：全类型 > 新能源 > 其他
        data: Dict[str, Dict[str, Any]] = {}
        for province, rows in candidates.items():
            chosen_row = self._select_best_row(rows)
            if chosen_row is None:
                continue

            i_val = ws.cell(row=chosen_row, column=9).value
            k_val = ws.cell(row=chosen_row, column=11).value
            m_val = ws.cell(row=chosen_row, column=13).value
            n_val = ws.cell(row=chosen_row, column=14).value
            l_val = ws.cell(row=chosen_row, column=12).value
            o_val = ws.cell(row=chosen_row, column=15).value

            data[province] = {
                "avg": self._to_float(i_val),
                "yoy": self._to_float(k_val),
                "wow": self._to_float(n_val),
                "prev_avg": self._to_float(m_val),
                "yoy_reason": self._to_str(l_val),
                "wow_reason": self._to_str(o_val),
            }

        # 按目标顺序排列地区
        ordered_regions = [r for r in SPOT_REGIONS if r in data]

        if len(ordered_regions) < len(SPOT_REGIONS):
            missing = [r for r in SPOT_REGIONS if r not in data]
            logger.warning("缺少现货数据的地区: %s", missing)

        return {
            "regions": ordered_regions,
            "data": data,
        }

    def _select_best_row(self, rows: List[Tuple[str, int]]) -> Optional[int]:
        """从候选行中按优先级选取最佳行。

        优先级：全类型 > 新能源 > 任意
        """
        # 优先：全类型
        for type_name, row_idx in rows:
            if type_name == "全类型":
                return row_idx
        # 其次：新能源
        for type_name, row_idx in rows:
            if type_name == "新能源":
                return row_idx
        # 最后：任意有效行
        return rows[0][1] if rows else None

    def _is_valid_spot_row(self, i_val: Any, k_val: Any) -> bool:
        """判断现货数据行是否有效。"""
        # I列必须有数值且非零
        if i_val is None:
            return False
        try:
            avg = float(i_val)
            if avg == 0:
                return False
        except (ValueError, TypeError):
            return False

        # K列不能是错误值
        if k_val is None:
            return False
        if isinstance(k_val, str):
            error_markers = ["#DIV/0!", "#VALUE!", "#N/A", "#REF!", "/"]
            if any(m in str(k_val) for m in error_markers):
                return False

        return True

    # ========================================================================
    # 原因描述文本采集
    # ========================================================================

    def _collect_reasons(self, wb: Any) -> Dict[str, Optional[str]]:
        """从国内数据填报表 H 列采集原因描述文本。

        Returns:
            {yoy_summary: "...", wow_summary: "...", ...}
        """
        sheet_name = "国内数据填报表"
        if sheet_name not in wb.sheetnames:
            self._errors.append({
                "level": "WARNING",
                "message": f"Sheet '{sheet_name}' 不存在，跳过原因数据",
            })
            return {}

        ws = wb[sheet_name]
        reasons: Dict[str, Optional[str]] = {}

        for row_num, field_name in REASON_ROWS:
            value = ws.cell(row=row_num, column=REASON_COL).value
            text = self._to_str(value)
            reasons[field_name] = text

            if text:
                logger.debug("原因 %s (R%d): %s", field_name, row_num, text[:50])
            else:
                logger.debug("原因 %s (R%d): 为空", field_name, row_num)

        return reasons

    # ========================================================================
    # 工具方法
    # ========================================================================

    def _to_float(self, value: Any) -> Optional[float]:
        """安全转换为浮点数。百分比值统一归一化到比率（如 -59.32% → -0.5932）。"""
        if value is None:
            return None
        try:
            if isinstance(value, (int, float)):
                return float(value)
            text = str(value).strip().replace(",", "").replace("，", "")
            if not text or text in ("-", "—", "N/A"):
                return None
            if "%" in text:
                text = text.replace("%", "").strip()
                # "59.32%" → 0.5932, "0.033" → 0.033
                return float(text) / 100.0
            return float(text)
        except (ValueError, TypeError):
            return None

    def _to_str(self, value: Any) -> Optional[str]:
        """安全转换为字符串。"""
        if value is None:
            return None
        text = str(value).strip()
        if not text or text in ("-", "—", "/", "N/A"):
            return None
        return text

    def _extract_year(self, filename: str) -> Optional[int]:
        """从文件名提取年份。"""
        match = re.search(r"(\d{4})年", filename)
        return int(match.group(1)) if match else None

    def _extract_week(self, filename: str) -> Optional[int]:
        """从文件名提取周数。"""
        match = re.search(r"第(\d+)周", filename)
        return int(match.group(1)) if match else None


def main() -> None:
    """命令行入口"""
    import argparse
    import json

    parser = argparse.ArgumentParser(description="汇总表补充数据采集")
    parser.add_argument("--input", required=True, help="汇总表.xlsx 路径")
    parser.add_argument("--output", help="JSON 输出路径")
    parser.add_argument("--year", type=int, help="年份")
    parser.add_argument("--week", type=int, help="周数")

    args = parser.parse_args()

    collector = SummaryCollector()
    result, errors = collector.collect(args.input, year=args.year, week=args.week)

    for err in errors:
        level = err.get("level", "INFO")
        msg = err.get("message", "")
        print(f"[{level}] {msg}")

    if result:
        # 现货摘要
        spot = result.get("spot_prices", {})
        regions = spot.get("regions", [])
        spot_data = spot.get("data", {})
        print(f"\n现货市场: {len(regions)}/{len(SPOT_REGIONS)} 个地区")
        for r in regions:
            d = spot_data.get(r, {})
            print(f"  {r}: 均价={d.get('avg')}, 同比={d.get('yoy')}, 环比={d.get('wow')}")

        # 原因摘要
        reasons = result.get("reasons", {})
        for key, val in reasons.items():
            preview = val[:60] if val else "(空)"
            print(f"\n原因 {key}: {preview}...")

    if args.output and result:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        with open(output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"\n数据已保存到: {args.output}")


if __name__ == "__main__":
    main()
