"""语义定位解析器 - 通过实体名称定位数据"""

from typing import Dict, List, Optional, Tuple
from openpyxl.worksheet.worksheet import Worksheet
from src.utils.entity_resolver import EntityResolver


class SemanticParser:
    """语义定位解析器，通过实体名称而非固定位置定位数据"""

    def __init__(self, resolver: EntityResolver):
        self.resolver = resolver

    def find_header_row(self, ws: Worksheet, max_rows: int = 20) -> Optional[int]:
        """
        找到表头行（包含能源类型或指标关键词的行）

        Args:
            ws: Worksheet 对象
            max_rows: 最大搜索行数

        Returns:
            表头行号（0-indexed），未找到返回 None
        """
        for row_idx in range(min(max_rows, ws.max_row)):
            row = ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1).__next__()
            header_count = 0
            for cell in row:
                text = str(cell.value) if cell.value else ""
                # 检查是否包含能源类型或指标关键词
                for keyword in self.resolver.get_all_energy_names():
                    if keyword in text:
                        header_count += 1
                for keyword in self.resolver.get_all_metric_names():
                    if keyword in text:
                        header_count += 1
            # 如果一行包含多个关键词，认为是表头
            if header_count >= 2:
                return row_idx
        return None

    def find_org_rows(self, ws: Worksheet, start_row: int = 0) -> Dict[str, int]:
        """
        找到所有组织所在行

        Args:
            ws: Worksheet 对象
            start_row: 起始行号

        Returns:
            {组织名: 行号} 映射（0-indexed）
        """
        org_rows = {}
        for row_idx in range(start_row, ws.max_row):
            row = ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1).__next__()
            for cell in row:
                text = str(cell.value) if cell.value else ""
                org_name, _ = self.resolver.resolve_organization(text)
                if org_name and org_name not in org_rows:
                    org_rows[org_name] = row_idx
        return org_rows

    def build_column_mapping(self, ws: Worksheet, header_row: int) -> Dict[Tuple[str, str], int]:
        """
        构建列映射（能源类型+指标 -> 列号）

        Args:
            ws: Worksheet 对象
            header_row: 表头行号（0-indexed）

        Returns:
            {(能源类型, 指标): 列号} 映射（0-indexed）
        """
        col_mapping = {}
        row = ws.iter_rows(min_row=header_row + 1, max_row=header_row + 1).__next__()

        for col_idx, cell in enumerate(row):
            text = str(cell.value) if cell.value else ""
            parsed = self.resolver.resolve_cell_header(text)

            energy_type = parsed.get("energy_type")
            metric = parsed.get("metric")

            if energy_type and metric:
                key = (energy_type, metric)
                col_mapping[key] = col_idx
            elif metric:
                # 仅指标，能源类型可能是合计
                # 检查是否已有能源类型在相邻列
                pass

        return col_mapping

    def parse_multi_level_header(
        self, ws: Worksheet, header_rows: List[int]
    ) -> Dict[Tuple[str, str], int]:
        """
        解析多层表头（如能源类型在上层，指标在下层）

        Args:
            ws: Worksheet 对象
            header_rows: 表头行号列表（如 [2, 3] 表示第3行是能源类型，第4行是指标）

        Returns:
            {(能源类型, 指标): 列号} 映射
        """
        col_mapping = {}

        # 找能源类型行和指标行
        energy_row = None
        metric_row = None

        for row_idx in header_rows:
            row = ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1).__next__()
            energy_count = 0
            metric_count = 0
            for cell in row:
                text = str(cell.value) if cell.value else ""
                for energy in self.resolver.get_all_energy_names():
                    if energy in text:
                        energy_count += 1
                for metric in self.resolver.get_all_metric_names():
                    if metric in text:
                        metric_count += 1

            if energy_count > metric_count:
                energy_row = row_idx
            elif metric_count > energy_count:
                metric_row = row_idx

        # 如果无法区分，假设第一行是能源，最后一行是指标
        if energy_row is None:
            energy_row = header_rows[0]
        if metric_row is None:
            metric_row = header_rows[-1]

        # 遍历列，找到能源类型并确定其范围
        current_energy = None

        for col_idx in range(ws.max_column):
            # 检查能源类型行
            cell = ws.cell(row=energy_row + 1, column=col_idx + 1)
            text = str(cell.value) if cell.value else ""

            # 尝试匹配能源类型
            et_name, _ = self.resolver.resolve_energy_type(text)
            if et_name:
                current_energy = et_name

            # 检查指标行
            cell = ws.cell(row=metric_row + 1, column=col_idx + 1)
            text = str(cell.value) if cell.value else ""

            # 尝试匹配指标
            metric_name, _ = self.resolver.resolve_metric(text)

            if current_energy and metric_name:
                col_mapping[(current_energy, metric_name)] = col_idx

        return col_mapping

    def extract_value(
        self, ws: Worksheet, row: int, col: int
    ) -> Tuple[Optional[float], Dict]:
        """
        提取单元格数值

        Args:
            ws: Worksheet 对象
            row: 行号（0-indexed）
            col: 列号（0-indexed）

        Returns:
            (数值, 源追溯信息)
        """
        cell = ws.cell(row=row + 1, column=col + 1)
        value = cell.value

        # 源追溯
        source_trace = {
            "row": row + 1,
            "col": col + 1,
            "cell": cell.coordinate,
        }

        # 数值转换
        if value is None:
            return None, source_trace

        if isinstance(value, (int, float)):
            return float(value), source_trace

        # 字符串转数值
        text = str(value).strip()
        try:
            # 处理带逗号的数字
            text = text.replace(",", "")
            # 处理百分号
            if "%" in text:
                text = text.replace("%", "")
                return float(text), source_trace
            return float(text), source_trace
        except ValueError:
            return None, source_trace

    def parse_sheet(
        self, ws: Worksheet, sheet_name: str = "unknown"
    ) -> Tuple[Dict, List[Dict]]:
        """
        解析整个 Sheet

        Args:
            ws: Worksheet 对象
            sheet_name: Sheet 名称

        Returns:
            (数据字典, 错误列表)
        """
        result = {"organizations": {}, "meta": {"sheet": sheet_name}}
        errors = []

        # 1. 找表头行（多层表头）
        # 尝试找多层表头：第3行是能源类型，第4行是指标
        header_row_1 = self.find_header_row(ws, max_rows=10)  # 能源类型行

        # 查找所有表头行（连续的关键词行）
        header_rows = []
        for row_idx in range(min(10, ws.max_row)):
            row = ws.iter_rows(min_row=row_idx + 1, max_row=row_idx + 1).__next__()
            header_count = 0
            for cell in row:
                text = str(cell.value) if cell.value else ""
                for keyword in self.resolver.get_all_energy_names() + self.resolver.get_all_metric_names():
                    if keyword in text:
                        header_count += 1
            if header_count >= 2:
                header_rows.append(row_idx)

        if not header_rows:
            errors.append({
                "level": "ERROR",
                "message": "未找到表头行",
                "sheet": sheet_name
            })
            return result, errors

        # 2. 找组织行
        data_start_row = max(header_rows) + 1
        org_rows = self.find_org_rows(ws, start_row=data_start_row)
        if not org_rows:
            errors.append({
                "level": "ERROR",
                "message": "未找到组织数据行",
                "sheet": sheet_name
            })
            return result, errors

        # 3. 构建列映射（多层表头）
        if len(header_rows) >= 2:
            col_mapping = self.parse_multi_level_header(ws, header_rows)
        else:
            col_mapping = self.build_column_mapping(ws, header_rows[0])

        # 4. 提取数据
        for org_name, row_idx in org_rows.items():
            org_data = {
                "name": org_name,
                "source": {"sheet": sheet_name, "row": row_idx + 1},
                "metrics": {}
            }

            # 获取组织信息
            _, org_info = self.resolver.resolve_organization(org_name)
            if org_info:
                org_data["id"] = org_info.get("id")
                org_data["full_name"] = org_info.get("full_name")
                org_data["category"] = org_info.get("category")
                org_data["region"] = org_info.get("region")

            # 提取各能源类型指标
            for (energy_type, metric), col_idx in col_mapping.items():
                value, source = self.extract_value(ws, row_idx, col_idx)
                source["sheet"] = sheet_name

                # 按能源类型分组
                if energy_type not in org_data["metrics"]:
                    org_data["metrics"][energy_type] = {}

                org_data["metrics"][energy_type][metric] = {
                    "value": value,
                    "source": source
                }

                # 记录缺失数据
                if value is None and energy_type != "合计":
                    errors.append({
                        "level": "WARN",
                        "message": f"{org_name}.{energy_type}.{metric} 数据缺失",
                        "field": f"{org_name}.{energy_type}.{metric}",
                        "sheet": sheet_name,
                        "cell": source.get("cell")
                    })

            result["organizations"][org_name] = org_data

        return result, errors


if __name__ == "__main__":
    # 测试（需要 Excel 文件）
    print("SemanticParser 已创建，需要 Excel 文件进行完整测试")
    print("用法: parser = SemanticParser(EntityResolver())")
    print("      data, errors = parser.parse_sheet(worksheet)")