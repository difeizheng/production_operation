"""综合分析报表采集器 - 从综合分析表直接读取数据

通过 CELL_MAP 中预定义的单元格映射，直接从综合分析报表.xlsx读取数据。
不依赖语义搜索，使用固定的单元格引用实现可靠采集。
"""

import logging
from pathlib import Path
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.collector.cell_map import (
    CELL_MAP,
    REPORT_TABLE_1_COLS,
    REPORT_TABLE_1_ROWS,
    CellRef,
    cell_coordinate,
    col_letter_to_idx,
)

logger = logging.getLogger(__name__)


class AnalysisCollector:
    """综合分析报表采集器

    从 综合分析报表.xlsx 的 综合分析表 sheet 读取预计算数据。
    """

    def __init__(self) -> None:
        self._cell_map = CELL_MAP
        self._errors: List[Dict[str, Any]] = []

    def collect(
        self,
        file_path: str,
        year: Optional[int] = None,
        week: Optional[int] = None,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        """从综合分析报表采集数据。

        Args:
            file_path: 综合分析报表.xlsx 文件路径
            year: 年份（用于元数据，可选）
            week: 周数（用于元数据，可选）

        Returns:
            (采集结果字典, 错误列表)
        """
        self._errors = []
        file_path = Path(file_path)

        if not file_path.exists():
            error = {"level": "ERROR", "message": f"文件不存在: {file_path}"}
            return {}, [error]

        # 加载工作簿
        try:
            wb = load_workbook(str(file_path), data_only=True)
        except Exception as e:
            error = {"level": "ERROR", "message": f"无法打开文件: {e}"}
            return {}, [error]

        # 获取综合分析表 sheet
        sheet_name = "综合分析表"
        if sheet_name not in wb.sheetnames:
            error = {"level": "ERROR", "message": f"Sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"}
            wb.close()
            return {}, [error]

        ws = wb[sheet_name]

        # 采集扁平数据
        flat_data = self._read_all_cells(ws)

        # 构建结构化数据
        result = self._build_structured_data(flat_data)

        # 构建报告表1
        result["report_table_1"] = self._build_report_table_1(ws)

        # 元数据
        result["meta"] = {
            "year": year or self._extract_year(file_path.name),
            "week": week or self._extract_week(file_path.name),
            "extracted_at": datetime.now().isoformat(),
            "source_file": file_path.name,
            "sheet_name": sheet_name,
            "collector": "AnalysisCollector",
        }

        # 验证报告
        result["validation_report"] = {
            "status": "pass" if not any(e["level"] == "ERROR" for e in self._errors) else "warning",
            "errors": self._errors,
            "field_count": len(flat_data),
            "coverage": self._calculate_coverage(flat_data),
        }

        wb.close()
        return result, self._errors

    def _read_all_cells(self, ws: Worksheet) -> Dict[str, Optional[float]]:
        """读取所有 CELL_MAP 中定义的单元格。

        Args:
            ws: Worksheet 对象

        Returns:
            {字段名: 数值} 字典
        """
        flat: Dict[str, Optional[float]] = {}

        for field_name, (col_letter, row) in self._cell_map.items():
            value = self._read_cell(ws, col_letter, row)
            flat[field_name] = value

            if value is None:
                logger.debug("字段 %s (%s%d) 为空", field_name, col_letter, row)

        logger.info("采集完成，共 %d 个字段", len(flat))
        return flat

    def _read_cell(
        self, ws: Worksheet, col_letter: str, row: int
    ) -> Optional[float]:
        """安全读取单元格数值。

        Args:
            ws: Worksheet 对象
            col_letter: 列字母（如 "I"）
            row: 行号（1-based）

        Returns:
            数值或 None
        """
        col_idx = col_letter_to_idx(col_letter)
        try:
            cell = ws.cell(row=row, column=col_idx)
            value = cell.value

            if value is None:
                return None

            if isinstance(value, (int, float)):
                return float(value)

            # 字符串转数值
            text = str(value).strip()
            if not text or text in ("-", "—", "N/A", "#DIV/0!", "#N/A", "#REF!"):
                return None

            text = text.replace(",", "").replace("，", "")
            if "%" in text:
                text = text.replace("%", "").strip()
                return float(text) / 100.0 if float(text) > 1 else float(text)

            return float(text)

        except (ValueError, TypeError) as e:
            coord = cell_coordinate(col_letter, row)
            logger.warning("无法读取 %s: %s", coord, e)
            return None

    def _build_structured_data(
        self, flat: Dict[str, Optional[float]]
    ) -> Dict[str, Any]:
        """将扁平数据构建为嵌套结构。

        Args:
            flat: 扁平字段数据

        Returns:
            嵌套结构数据
        """
        result: Dict[str, Any] = {
            "domestic": self._build_domestic(flat),
            "international": self._build_international(flat),
            "organizations": {},  # 保留兼容性
        }
        return result

    def _build_domestic(
        self, flat: Dict[str, Optional[float]]
    ) -> Dict[str, Any]:
        """构建国内数据分析结构。"""
        return {
            "electricity": {
                "total": flat.get("dom.electricity.total"),
                "hydro": flat.get("dom.electricity.hydro"),
                "new_energy": flat.get("dom.electricity.new_energy"),
                "wind": flat.get("dom.electricity.wind"),
                "solar": flat.get("dom.electricity.solar"),
                "thermal": flat.get("dom.electricity.thermal"),
            },
            "price": {
                "total": flat.get("dom.price.total"),
                "hydro": flat.get("dom.price.hydro"),
                "new_energy": flat.get("dom.price.new_energy"),
                "wind": flat.get("dom.price.wind"),
                "solar": flat.get("dom.price.solar"),
                "thermal": flat.get("dom.price.thermal"),
            },
            "revenue": {
                "total": flat.get("dom.revenue.total"),
                "hydro": flat.get("dom.revenue.hydro"),
                "new_energy": flat.get("dom.revenue.new_energy"),
                "wind": flat.get("dom.revenue.wind"),
                "solar": flat.get("dom.revenue.solar"),
                "thermal": flat.get("dom.revenue.thermal"),
            },
            "yoy": {
                "electricity": flat.get("dom.yoy.electricity.total"),
                "revenue": flat.get("report.yoy_revenue.total"),
                "price_change": flat.get("dom.yoy.price_change.total"),
                "price_impact": flat.get("dom.yoy.price_impact.total"),
                "share_impact": flat.get("dom.yoy.share_impact.total"),
                "hydro": {
                    "price_change": flat.get("dom.yoy.price_change.hydro"),
                    "share_change": flat.get("dom.yoy.share_change.hydro"),
                    "combined_impact": flat.get("dom.yoy.combined_impact.hydro"),
                },
                "new_energy": {
                    "price_change": flat.get("dom.yoy.price_change.new_energy"),
                    "share_change": flat.get("dom.yoy.share_change.new_energy"),
                    "combined_impact": flat.get("dom.yoy.combined_impact.new_energy"),
                },
                "wind": {
                    "price_change": flat.get("dom.yoy.price_change.wind"),
                    "share_change": flat.get("dom.yoy.share_change.wind"),
                    "combined_impact": flat.get("dom.yoy.combined_impact.wind"),
                },
                "solar": {
                    "price_change": flat.get("dom.yoy.price_change.solar"),
                    "share_change": flat.get("dom.yoy.share_change.solar"),
                    "combined_impact": flat.get("dom.yoy.combined_impact.solar"),
                },
                "thermal": {
                    "price_change": flat.get("dom.yoy.price_change.thermal"),
                    "share_change": flat.get("dom.yoy.share_change.thermal"),
                    "combined_impact": flat.get("dom.yoy.combined_impact.thermal"),
                },
            },
            "wow": {
                "electricity": flat.get("dom.wow.electricity.total"),
                "revenue": flat.get("report.wow_revenue.total"),
                "price_change": flat.get("dom.wow.price_change.total"),
                "price_impact": flat.get("dom.wow.price_impact.total"),
                "share_impact": flat.get("dom.wow.share_impact.total"),
                "hydro": {
                    "price_change": flat.get("dom.wow.price_change.hydro"),
                    "share_change": flat.get("dom.wow.share_change.hydro"),
                    "combined_impact": flat.get("dom.wow.combined_impact.hydro"),
                },
                "new_energy": {
                    "price_change": flat.get("dom.wow.price_change.new_energy"),
                    "share_change": flat.get("dom.wow.share_change.new_energy"),
                    "combined_impact": flat.get("dom.wow.combined_impact.new_energy"),
                },
                "wind": {
                    "price_change": flat.get("dom.wow.price_change.wind"),
                    "share_change": flat.get("dom.wow.share_change.wind"),
                    "combined_impact": flat.get("dom.wow.combined_impact.wind"),
                },
                "solar": {
                    "price_change": flat.get("dom.wow.price_change.solar"),
                    "share_change": flat.get("dom.wow.share_change.solar"),
                    "combined_impact": flat.get("dom.wow.combined_impact.solar"),
                },
                "thermal": {
                    "price_change": flat.get("dom.wow.price_change.thermal"),
                    "share_change": flat.get("dom.wow.share_change.thermal"),
                    "combined_impact": flat.get("dom.wow.combined_impact.thermal"),
                },
            },
        }

    def _build_international(
        self, flat: Dict[str, Optional[float]]
    ) -> Dict[str, Any]:
        """构建国际数据分析结构。"""
        return {
            "electricity": {
                "total": flat.get("intl.electricity.total"),
                "hydro": flat.get("intl.electricity.hydro"),
                "new_energy": flat.get("intl.electricity.new_energy"),
            },
            "price": {
                "total": flat.get("intl.price.total"),
            },
            "revenue": {
                "total": flat.get("intl.revenue.total"),
            },
            "yoy": {
                "electricity": flat.get("intl.yoy.electricity.total"),
                "price_change": flat.get("intl.yoy.price_change.total"),
            },
            "wow": {
                "electricity": flat.get("intl.wow.electricity.total"),
                "price_change": flat.get("intl.wow.price_change.total"),
            },
        }

    def _build_report_table_1(self, ws: Worksheet) -> Dict[str, Any]:
        """构建报告表1数据（已转换单位）。

        从 Rows 76-86 读取，生成与真实文档 Table 1 一致的结构。

        Returns:
            {headers: [...], rows: [[metric, hydro, new_energy, wind, solar, thermal, total], ...]}
        """
        headers = ["", "水电", "新能源", "风电", "光伏", "火电", "合计"]
        col_letters = ["C", "D", "E", "F", "G", "I"]

        row_labels = [
            (78, "国内上网电量"),
            (79, "同比"),
            (80, "环比"),
            (81, "国内上网电价"),
            (82, "同比"),
            (83, "环比"),
            (84, "国内发电收入"),
            (85, "同比"),
            (86, "环比"),
        ]

        rows: List[List[Optional[float]]] = []
        for row_num, label in row_labels:
            row_data: List[Optional[float]] = []
            for col_letter in col_letters:
                row_data.append(self._read_cell(ws, col_letter, row_num))
            rows.append(row_data)

        return {
            "headers": headers,
            "row_labels": [label for _, label in row_labels],
            "row_numbers": [row_num for row_num, _ in row_labels],
            "data": rows,
        }

    def _calculate_coverage(self, flat: Dict[str, Optional[float]]) -> float:
        """计算数据覆盖率。"""
        total = len(flat)
        if total == 0:
            return 0.0
        filled = sum(1 for v in flat.values() if v is not None)
        return round(filled / total * 100, 2)

    def _extract_year(self, filename: str) -> Optional[int]:
        """从文件名提取年份。"""
        import re
        match = re.search(r"(\d{4})年", filename)
        return int(match.group(1)) if match else None

    def _extract_week(self, filename: str) -> Optional[int]:
        """从文件名提取周数。"""
        import re
        match = re.search(r"第(\d+)周", filename)
        return int(match.group(1)) if match else None


def main() -> None:
    """命令行入口"""
    import argparse
    import json

    parser = argparse.ArgumentParser(description="综合分析报表数据采集")
    parser.add_argument("--input", required=True, help="综合分析报表.xlsx 路径")
    parser.add_argument("--output", help="JSON 输出路径")
    parser.add_argument("--year", type=int, help="年份")
    parser.add_argument("--week", type=int, help="周数")

    args = parser.parse_args()

    collector = AnalysisCollector()
    result, errors = collector.collect(args.input, year=args.year, week=args.week)

    # 输出错误
    for err in errors:
        level = err.get("level", "INFO")
        msg = err.get("message", "")
        print(f"[{level}] {msg}")

    # 输出摘要
    if result:
        meta = result.get("meta", {})
        validation = result.get("validation_report", {})
        print(f"\n年份: {meta.get('year')}, 周数: {meta.get('week')}")
        print(f"字段数: {validation.get('field_count', 0)}")
        print(f"覆盖率: {validation.get('coverage', 0)}%")

        # 输出报告表1预览
        table = result.get("report_table_1", {})
        print(f"\n报告表1 ({len(table.get('data', []))} rows):")
        for i, row in enumerate(table.get("data", [])):
            label = table.get("row_labels", [])[i] if i < len(table.get("row_labels", [])) else ""
            formatted = [f"{v:.4f}" if v is not None else "-" for v in row]
            print(f"  {label}: {formatted}")

    # 保存
    if args.output and result:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        with open(output, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        print(f"\n数据已保存到: {args.output}")


if __name__ == "__main__":
    main()
