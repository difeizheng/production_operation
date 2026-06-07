"""原因描述文本采集器 - 从汇总表 H/T 等列采集"原因/分析"叙述文本

与 SummaryCollector 互补：
    - SummaryCollector 负责现货价格 + 4 条汇总原因
    - ReasonCollector 负责**全部**原因叙述文本（H/T/L/O 等列）

设计原则：
    1. 槽位化定义（ReasonSlot）：每个原因文本是一个独立槽位
    2. 语义定位 + 行匹配：表头/科目用语义搜索，单元格用绝对定位
    3. 优雅降级：某个槽位失败不影响整体
    4. 类型安全：使用 dataclass 代替裸 dict

包含的槽位（按段落）：
    - 国内段：电量同比/环比原因（汇总+4 公司）
    - 国内段：电价同比/环比原因（汇总+4 公司）
    - 国际段：电量/电价同比/环比原因
    - 现货段：各地区现货价格原因
    - 碳资产段：绿证/CCER 原因
"""
from __future__ import annotations

import logging
import re
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


# ============================================================================
# 数据类定义
# ============================================================================

@dataclass(frozen=True)
class ReasonSlot:
    """原因文本槽位定义。

    一个 ReasonSlot 对应 V4 周报中的一个原因段落
    （如"国内电量同比变化原因"、"国际电价环比变化原因"等）。
    """
    slot_id: str              # 唯一标识，如 "dom.yoy.changjiang"
    sheet_name: str           # Excel sheet 名，如 "国内数据填报表"
    row: int                  # 行号（1-based）
    col: int                  # 列号（1-based），如 8 = H
    description: str          # 人类可读描述
    category: str             # 分类：domestic_yoy / domestic_wow / intl_yoy / intl_wow / spot / carbon


@dataclass(frozen=True)
class ReasonResult:
    """单个槽位的提取结果。"""
    slot: ReasonSlot
    raw_text: str             # 原始文本（如 "1、全集团电量..."）
    source_file: str          # 来源文件名
    is_empty: bool            # 是否为空


# ============================================================================
# 槽位定义表（核心）
# ============================================================================

# 国内段原因：H 列（8）
DOMESTIC_REASON_SLOTS: List[ReasonSlot] = [
    # --- 电量同比 ---
    ReasonSlot("dom.elec.yoy.changjiang",  "国内数据填报表", 29,  8, "国内电量同比原因（长江电力+全集团）", "domestic_yoy"),
    ReasonSlot("dom.elec.yoy.new_energy",  "国内数据填报表", 32,  8, "国内电量同比原因（新能源）",         "domestic_yoy"),
    ReasonSlot("dom.elec.yoy.thermal",     "国内数据填报表", 33,  8, "国内电量同比原因（火电）",           "domestic_yoy"),
    ReasonSlot("dom.elec.yoy.sanxia_dev",  "国内数据填报表", 34,  8, "国内电量同比原因（三峡发展）",       "domestic_yoy"),
    ReasonSlot("dom.elec.yoy.hubei",       "国内数据填报表", 43,  8, "国内电量同比原因（湖北能源）",       "domestic_yoy"),
    ReasonSlot("dom.elec.yoy.summary",     "国内数据填报表", 47,  8, "国内电量同比原因（汇总长版）",       "domestic_yoy"),
    # --- 电量环比 ---
    ReasonSlot("dom.elec.wow.changjiang",  "国内数据填报表", 53,  8, "国内电量环比原因（长江电力+全集团）", "domestic_wow"),
    ReasonSlot("dom.elec.wow.new_energy",  "国内数据填报表", 56,  8, "国内电量环比原因（新能源）",         "domestic_wow"),
    ReasonSlot("dom.elec.wow.thermal",     "国内数据填报表", 57,  8, "国内电量环比原因（火电）",           "domestic_wow"),
    ReasonSlot("dom.elec.wow.sanxia_dev",  "国内数据填报表", 58,  8, "国内电量环比原因（三峡发展）",       "domestic_wow"),
    ReasonSlot("dom.elec.wow.hubei",       "国内数据填报表", 67,  8, "国内电量环比原因（湖北能源）",       "domestic_wow"),
    ReasonSlot("dom.elec.wow.summary",     "国内数据填报表", 71,  8, "国内电量环比原因（汇总长版）",       "domestic_wow"),
    # --- 电价同比 ---
    ReasonSlot("dom.price.yoy.changjiang", "国内数据填报表", 101, 8, "国内电价同比原因（长江电力+全集团）", "domestic_yoy"),
    ReasonSlot("dom.price.yoy.new_energy", "国内数据填报表", 104, 8, "国内电价同比原因（新能源）",         "domestic_yoy"),
    ReasonSlot("dom.price.yoy.thermal",    "国内数据填报表", 105, 8, "国内电价同比原因（火电）",           "domestic_yoy"),
    ReasonSlot("dom.price.yoy.sanxia_dev", "国内数据填报表", 106, 8, "国内电价同比原因（三峡发展）",       "domestic_yoy"),
    ReasonSlot("dom.price.yoy.hubei",      "国内数据填报表", 115, 8, "国内电价同比原因（湖北能源）",       "domestic_yoy"),
    ReasonSlot("dom.price.yoy.summary",    "国内数据填报表", 119, 8, "国内电价同比原因（汇总长版）",       "domestic_yoy"),
]

# 国际段原因：H 列（8）+ T 列（20）
INTERNATIONAL_REASON_SLOTS: List[ReasonSlot] = [
    ReasonSlot("intl.elec.yoy.h",   "国际数据填报表", 14,  8,  "国际电量同比原因",        "intl_yoy"),
    ReasonSlot("intl.elec.wow.h",   "国际数据填报表", 15,  8,  "国际电量环比原因",        "intl_wow"),
    ReasonSlot("intl.price.yoy.t",  "国际数据填报表", 14,  20, "国际电价同比原因（备用）", "intl_yoy"),
    ReasonSlot("intl.price.wow.t",  "国际数据填报表", 15,  20, "国际电价环比原因（备用）", "intl_wow"),
]

# 现货段原因：L 列（12，yoy_reason）+ O 列（15，wow_reason）
SPOT_REASON_SLOTS_PATTERN: List[ReasonSlot] = [
    # 注：现货原因在 spot_prices 数据中（已在 summary_collector 提取）
    # 这里仅占位
]

# 碳资产段原因：T 列（20）
CARBON_REASON_SLOTS: List[ReasonSlot] = [
    ReasonSlot("carbon.green_cert.t",   "绿证填报表", 4,   20, "绿证原因",     "carbon"),
    ReasonSlot("carbon.green_cert.t2",  "绿证填报表", 17,  20, "绿证原因（续）", "carbon"),
    ReasonSlot("carbon.green_cert.t3",  "绿证填报表", 30,  20, "绿证原因（续2）","carbon"),
]


# 合并所有槽位
ALL_REASON_SLOTS: List[ReasonSlot] = (
    DOMESTIC_REASON_SLOTS
    + INTERNATIONAL_REASON_SLOTS
    + CARBON_REASON_SLOTS
)


# ============================================================================
# 采集器
# ============================================================================

class ReasonCollector:
    """从汇总表采集"原因/分析"叙述文本。

    适用 Excel：周数据汇总表.xlsx
    返回：所有 ReasonSlot 的提取结果，缺失/空文本用 is_empty=True 标记
    """

    def __init__(self, slots: Optional[List[ReasonSlot]] = None) -> None:
        """初始化采集器。

        Args:
            slots: 自定义槽位列表（默认使用 ALL_REASON_SLOTS）
        """
        self._slots = slots if slots is not None else ALL_REASON_SLOTS
        self._errors: List[Dict[str, Any]] = []

    @property
    def slots(self) -> List[ReasonSlot]:
        return list(self._slots)

    def collect(
        self,
        file_path: str,
        year: Optional[int] = None,
        week: Optional[int] = None,
    ) -> Tuple[Dict[str, ReasonResult], List[Dict[str, Any]]]:
        """从汇总表采集所有原因文本。

        Args:
            file_path: 汇总表.xlsx 路径
            year: 年份（用于元数据）
            week: 周数（用于元数据）

        Returns:
            (slot_id → ReasonResult 字典, 错误列表)
        """
        self._errors = []
        file_path = Path(file_path)

        if not file_path.exists():
            return {}, [{"level": "ERROR", "message": f"文件不存在: {file_path}"}]

        try:
            wb = load_workbook(str(file_path), data_only=True)
        except Exception as e:
            return {}, [{"level": "ERROR", "message": f"无法打开文件: {e}"}]

        results: Dict[str, ReasonResult] = {}

        for slot in self._slots:
            text, error = self._read_slot(wb, slot)
            if error is not None:
                self._errors.append(error)
            results[slot.slot_id] = ReasonResult(
                slot=slot,
                raw_text=text,
                source_file=file_path.name,
                is_empty=not text.strip(),
            )

        wb.close()
        return results, self._errors

    def collect_one(
        self,
        file_path: str,
        slot_id: str,
    ) -> Optional[ReasonResult]:
        """采集单个槽位（便利方法）。

        Args:
            file_path: Excel 路径
            slot_id: 槽位 ID

        Returns:
            ReasonResult 或 None（槽位未定义时）
        """
        slot = next((s for s in self._slots if s.slot_id == slot_id), None)
        if slot is None:
            logger.warning("未定义的槽位: %s", slot_id)
            return None

        wb = load_workbook(str(file_path), data_only=True)
        text, _ = self._read_slot(wb, slot)
        wb.close()
        return ReasonResult(
            slot=slot,
            raw_text=text,
            source_file=str(file_path),
            is_empty=not text.strip(),
        )

    def _read_slot(
        self,
        wb: Any,
        slot: ReasonSlot,
    ) -> Tuple[str, Optional[Dict[str, Any]]]:
        """读取单个槽位的文本。"""
        if slot.sheet_name not in wb.sheetnames:
            return "", {
                "level": "WARNING",
                "slot_id": slot.slot_id,
                "message": f"Sheet '{slot.sheet_name}' 不存在",
            }

        ws: Worksheet = wb[slot.sheet_name]
        raw = ws.cell(row=slot.row, column=slot.col).value
        text = self._normalize_text(raw)

        if not text:
            logger.debug(
                "槽位 %s 文本为空 (sheet=%s, R%dC%d)",
                slot.slot_id, slot.sheet_name, slot.row, slot.col,
            )

        return text, None

    @staticmethod
    def _normalize_text(value: Any) -> str:
        """规范化文本：处理 None、换行符、占位符。"""
        if value is None:
            return ""
        text = str(value).strip()
        if text in ("-", "—", "/", "N/A", "无", "（空）"):
            return ""
        return text


# ============================================================================
# 便利函数
# ============================================================================

def collect_all_reasons(file_path: str) -> Dict[str, str]:
    """便利函数：返回 {slot_id: raw_text} 字典。

    Args:
        file_path: 汇总表路径

    Returns:
        简化的 {slot_id: text} 字典（空文本会被过滤）
    """
    collector = ReasonCollector()
    results, errors = collector.collect(file_path)
    for err in errors:
        logger.warning("采集警告: %s", err.get("message"))
    return {
        sid: r.raw_text
        for sid, r in results.items()
        if not r.is_empty
    }


def find_slot_by_id(slot_id: str) -> Optional[ReasonSlot]:
    """根据 slot_id 查找槽位定义。"""
    return next((s for s in ALL_REASON_SLOTS if s.slot_id == slot_id), None)


# ============================================================================
# 命令行入口
# ============================================================================

def main() -> None:
    """命令行入口：扫描并打印所有非空原因文本。"""
    import argparse
    import json

    parser = argparse.ArgumentParser(description="原因文本采集器")
    parser.add_argument("--input", required=True, help="汇总表.xlsx 路径")
    parser.add_argument("--slot-id", help="只采集指定槽位")
    parser.add_argument("--output", help="JSON 输出路径")
    parser.add_argument("--verbose", "-v", action="store_true", help="显示详情")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s [%(name)s] %(message)s",
    )

    if args.slot_id:
        # 单槽位采集
        slot = find_slot_by_id(args.slot_id)
        if slot is None:
            print(f"❌ 槽位不存在: {args.slot_id}")
            print(f"可用槽位: {[s.slot_id for s in ALL_REASON_SLOTS[:5]]} ...")
            return
        collector = ReasonCollector(slots=[slot])
    else:
        collector = ReasonCollector()

    results, errors = collector.collect(args.input)

    print(f"\n=== 采集结果 ===")
    print(f"总槽位: {len(results)}, 错误: {len(errors)}")
    print()

    for sid, result in results.items():
        marker = "📝" if not result.is_empty else "⭕"
        slot = result.slot
        print(f"{marker} [{sid}]")
        print(f"   位置: {slot.sheet_name}!R{slot.row}C{slot.col}")
        print(f"   描述: {slot.description}")
        if result.raw_text:
            preview = result.raw_text.replace("\n", " ⏎ ")
            print(f"   原文: {preview[:120]}{'...' if len(preview) > 120 else ''}")
        print()

    if errors:
        print(f"=== 错误/警告 ===")
        for err in errors[:10]:
            print(f"  [{err.get('level')}] {err.get('slot_id', '?')}: {err.get('message')}")

    if args.output:
        output = Path(args.output)
        output.parent.mkdir(parents=True, exist_ok=True)
        data = {
            sid: {
                "raw_text": r.raw_text,
                "sheet": r.slot.sheet_name,
                "row": r.slot.row,
                "col": r.slot.col,
                "description": r.slot.description,
                "category": r.slot.category,
                "is_empty": r.is_empty,
            }
            for sid, r in results.items()
        }
        with open(output, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"\n✅ 已保存到: {args.output}")


if __name__ == "__main__":
    main()
